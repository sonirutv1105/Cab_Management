import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fpdf import FPDF

class ExportPDF(FPDF):
    def __init__(self, title="Export", *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.doc_title = title

    def header(self):
        # Arial bold 15
        self.set_font('helvetica', 'B', 16)
        # Move to the right
        self.cell(80)
        # Title
        self.cell(30, 10, 'CMS Enterprise Data Export', 0, 1, 'C')
        
        self.set_font('helvetica', 'I', 12)
        self.cell(0, 10, f'{self.doc_title}', 0, 1, 'C')
        
        self.set_font('helvetica', '', 9)
        self.cell(0, 8, f'Generated at: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'R')
        self.ln(5)

    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        # Arial italic 8
        self.set_font('helvetica', 'I', 8)
        # Page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')

def generate_pdf(title: str, headers: list, rows: list) -> io.BytesIO:
    # Use landscape for wide tables
    pdf = ExportPDF(title=title, orientation='L', format='A4')
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("helvetica", size=8)

    # Calculate column widths
    epw = pdf.w - 2 * pdf.l_margin
    col_width = epw / len(headers) if headers else epw
    
    # Table Header
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(200, 220, 255)
    
    # Calculate heights
    th = pdf.font_size * 2

    # Draw header
    for header in headers:
        pdf.cell(col_width, th, str(header)[:30], border=1, align='C', fill=True)
    pdf.ln(th)

    # Table Body
    pdf.set_font("helvetica", "", 8)
    for row in rows:
        # Check if page break is needed
        if pdf.get_y() + th > pdf.page_break_trigger:
            pdf.add_page()
            # Redraw header
            pdf.set_font("helvetica", "B", 9)
            for header in headers:
                pdf.cell(col_width, th, str(header)[:30], border=1, align='C', fill=True)
            pdf.ln(th)
            pdf.set_font("helvetica", "", 8)

        # Draw row
        for item in row:
            # truncate long strings so they don't break the cell
            text = str(item) if item is not None else ""
            text = (text[:40] + '..') if len(text) > 40 else text
            pdf.cell(col_width, th, text, border=1, align='L')
        pdf.ln(th)

    pdf_buffer = io.BytesIO()
    pdf_out = pdf.output(dest='S')
    # fpdf's string output must be encoded to bytes for BytesIO
    pdf_buffer.write(pdf_out.encode('latin-1'))
    pdf_buffer.seek(0)
    return pdf_buffer

def generate_excel(title: str, headers: list, rows: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] # Excel limits sheet name to 31 chars

    # Add Title
    ws.append([f"CMS Enterprise Data Export - {title}"])
    ws.append([f"Generated at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([]) # blank row
    
    # Add headers
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[4]: # Row 4 is where headers are
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Add rows
    for row in rows:
        ws.append(row)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
